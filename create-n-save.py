import openpyxl
from openpyxl import Workbook

''' По подразбиране openpyxl когато запаметява файл, който не сме отворили
чрез load_workbook, направо презаписва самият файл! '''

wb = openpyxl.load_workbook('test.xlsx')

workbook = Workbook()
workbook.save(filename="test.xlsx")
wb = openpyxl.Workbook()
sheet = wb.active

x1 = sheet.cell(row = 1, column = 1)
x1.value = "Proba1"

x2 = sheet.cell(row = 1, column = 2)
x2.value = "Proba2"

x3 = sheet['A2']
x3.value = 'demo1'

x4 = sheet['B2']
x4.value = 'demo2'

x5 = sheet['C2']
x5.value = 150

wb.save('test.xlsx')